"""services/payroll_export.py — E11 slip gaji driver (PDF reportlab + Excel openpyxl).

Menerima dict payout (dari koleksi driver_payouts) + optional company_info. Tidak menyentuh DB.
"""
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PERIOD_LABEL = {"monthly": "Bulanan", "weekly": "Mingguan", "per_trip": "Per Trip"}
STATUS_LABEL = {"draft": "Draft", "approved": "Disetujui", "paid": "Dibayar"}


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


def _grid(rows, widths, aligns=None):
    t = Table(rows, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")),
    ]
    for col in (aligns or []):
        style.append(("ALIGN", (col, 0), (col, -1), "RIGHT"))
    t.setStyle(TableStyle(style))
    return t


def _components(p: dict):
    rows = [["Komponen", "Nilai"]]
    rows.append([f"Gaji Pokok ({PERIOD_LABEL.get(p.get('period_type'), '-')})", _rp(p.get("base_salary"))])
    rows.append([f"Komisi per Trip ({p.get('trips_count', 0)} trip)", _rp(p.get("commission_trip"))])
    rows.append([f"Komisi % Revenue ({_rp(p.get('total_revenue'))})", _rp(p.get("commission_pct"))])
    rows.append([f"Uang Jalan ({p.get('total_km', 0)} km)", _rp(p.get("allowance_km"))])
    rows.append(["Subtotal (Gross)", _rp(p.get("gross"))])
    for b in p.get("bonuses") or []:
        rows.append([f"Bonus · {b.get('label') or 'Bonus'}", _rp(b.get("amount"))])
    for d in p.get("deductions") or []:
        rows.append([f"Potongan · {d.get('label') or 'Potongan'}", f"-{_rp(d.get('amount'))}"])
    return rows


# --------------------------------------------------------------------------- PDF
def slip_pdf(payout: dict, company: dict = None) -> bytes:
    company = company or {}
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title="Slip Gaji Driver")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=18, spaceAfter=2)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#6B6B73"))
    big = ParagraphStyle("big", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#0058CC"))
    el = []
    el.append(Paragraph(company.get("name") or "Slip Gaji Driver", h1))
    el.append(Paragraph("Slip Pembayaran / Payroll Driver", small))
    el.append(Spacer(1, 10))

    info = [
        ["Driver", str(payout.get("driver_name") or "-")],
        ["Periode", f"{payout.get('period_start')} — {payout.get('period_end')} ({PERIOD_LABEL.get(payout.get('period_type'), '-')})"],
        ["Status", STATUS_LABEL.get(payout.get("status"), payout.get("status"))],
        ["Trip Selesai", str(payout.get("trips_count", 0))],
        ["Total KM", f"{payout.get('total_km', 0)} km"],
    ]
    el.append(_grid([["Info", "Detail"]] + info, [55 * mm, 115 * mm]))
    el.append(Spacer(1, 12))

    el.append(_grid(_components(payout), [110 * mm, 60 * mm], aligns=[1]))
    el.append(Spacer(1, 12))

    el.append(Paragraph(f"TOTAL DITERIMA: {_rp(payout.get('total'))}", big))
    if payout.get("notes"):
        el.append(Spacer(1, 8))
        el.append(Paragraph(f"Catatan: {payout.get('notes')}", small))
    doc.build(el)
    return buf.getvalue()


# --------------------------------------------------------------------------- Excel
def slip_xlsx(payout: dict, company: dict = None) -> bytes:
    company = company or {}
    wb = Workbook()
    hfill = PatternFill("solid", fgColor="1C1C1E")
    hfont = Font(color="FFFFFF", bold=True)

    def head(ws, n):
        for c in range(1, n + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 42)

    ws = wb.active
    ws.title = "Slip Gaji"
    ws.append(["Info", "Detail"])
    head(ws, 2)
    ws.append(["Perusahaan", company.get("name") or "-"])
    ws.append(["Driver", payout.get("driver_name")])
    ws.append(["Periode", f"{payout.get('period_start')} — {payout.get('period_end')}"])
    ws.append(["Tipe Periode", PERIOD_LABEL.get(payout.get("period_type"), "-")])
    ws.append(["Status", STATUS_LABEL.get(payout.get("status"), payout.get("status"))])
    ws.append(["Trip Selesai", payout.get("trips_count", 0)])
    ws.append(["Total KM", payout.get("total_km", 0)])
    ws.append([])

    ws2 = wb.create_sheet("Rincian")
    ws2.append(["Komponen", "Nilai"])
    head(ws2, 2)
    ws2.append([f"Gaji Pokok ({PERIOD_LABEL.get(payout.get('period_type'), '-')})", payout.get("base_salary")])
    ws2.append([f"Komisi per Trip ({payout.get('trips_count', 0)} trip)", payout.get("commission_trip")])
    ws2.append([f"Komisi % Revenue", payout.get("commission_pct")])
    ws2.append([f"Uang Jalan ({payout.get('total_km', 0)} km)", payout.get("allowance_km")])
    ws2.append(["Subtotal (Gross)", payout.get("gross")])
    for b in payout.get("bonuses") or []:
        ws2.append([f"Bonus · {b.get('label') or 'Bonus'}", b.get("amount")])
    for d in payout.get("deductions") or []:
        ws2.append([f"Potongan · {d.get('label') or 'Potongan'}", -abs(float(d.get("amount") or 0))])
    ws2.append(["TOTAL DITERIMA", payout.get("total")])
    autosize(ws)
    autosize(ws2)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
