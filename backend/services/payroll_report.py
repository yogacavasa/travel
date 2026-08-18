"""services/payroll_report.py — Rekap Payroll untuk modul Laporan (E12).

Agregasi read-only dari koleksi driver_payouts + expenses (gaji_driver) + payments:
  - Ringkasan periode (count per status, total gross/bonus/potongan/net)
  - Rincian per-driver pada periode + kolom YTD (akrual tahun berjalan)
  - Biaya SDM vs revenue (rasio payroll_cost / revenue periode)
Export PDF (reportlab) + Excel (openpyxl).
"""
from datetime import datetime, timezone
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

STATUS_LABEL = {"draft": "Draft", "approved": "Disetujui", "paid": "Dibayar"}


def _month(s):
    return str(s or "")[:7]


async def build(db, period=None) -> dict:
    period = period or datetime.now(timezone.utc).strftime("%Y-%m")
    year = period[:4]
    payouts = await db.driver_payouts.find({}, {"_id": 0}).to_list(10000)

    in_period = [p for p in payouts if _month(p.get("period_start")) == period]
    by_status = {"draft": 0, "approved": 0, "paid": 0}
    tot_gross = tot_bonus = tot_ded = tot_net = 0.0
    per_driver = {}
    for p in in_period:
        st = p.get("status", "draft")
        by_status[st] = by_status.get(st, 0) + 1
        tot_gross += float(p.get("gross") or 0)
        tot_bonus += float(p.get("bonus_total") or 0)
        tot_ded += float(p.get("deduction_total") or 0)
        tot_net += float(p.get("total") or 0)
        d = per_driver.setdefault(p.get("driver_id"), {
            "driver_id": p.get("driver_id"), "driver_name": p.get("driver_name"),
            "trips": 0, "km": 0.0, "gross": 0.0, "bonus": 0.0, "deduction": 0.0,
            "total": 0.0, "statuses": [], "ytd": 0.0,
        })
        d["trips"] += int(p.get("trips_count") or 0)
        d["km"] += float(p.get("total_km") or 0)
        d["gross"] += float(p.get("gross") or 0)
        d["bonus"] += float(p.get("bonus_total") or 0)
        d["deduction"] += float(p.get("deduction_total") or 0)
        d["total"] += float(p.get("total") or 0)
        if STATUS_LABEL.get(st, st) not in d["statuses"]:
            d["statuses"].append(STATUS_LABEL.get(st, st))

    # YTD (akrual tahun berjalan: approved + paid) per driver
    for p in payouts:
        if _month(p.get("period_start"))[:4] == year and p.get("status") in ("approved", "paid"):
            row = per_driver.get(p.get("driver_id"))
            if row is not None:
                row["ytd"] += float(p.get("total") or 0)

    driver_rows = sorted(per_driver.values(), key=lambda r: r["total"], reverse=True)
    for r in driver_rows:
        r["km"] = round(r["km"], 1)
        for k in ("gross", "bonus", "deduction", "total", "ytd"):
            r[k] = round(r[k], 2)

    # Biaya SDM vs revenue (periode)
    payroll_cost_rows = await db.expenses.aggregate([
        {"$match": {"category": "gaji_driver"}},
        {"$group": {"_id": {"$substrBytes": [{"$ifNull": ["$created_at", ""]}, 0, 7]},
                    "t": {"$sum": {"$ifNull": ["$amount", 0]}}}},
    ]).to_list(1000)
    payroll_cost = round(sum(float(r["t"]) for r in payroll_cost_rows if r["_id"] == period), 2)
    rev_rows = await db.payments.aggregate([
        {"$group": {"_id": {"$substrBytes": [{"$ifNull": ["$paid_at", ""]}, 0, 7]},
                    "t": {"$sum": {"$ifNull": ["$amount", 0]}}}},
    ]).to_list(1000)
    revenue = round(sum(float(r["t"]) for r in rev_rows if r["_id"] == period), 2)
    ratio = round(payroll_cost / revenue * 100, 1) if revenue > 0 else 0.0

    return {
        "period": period,
        "count": len(in_period),
        "by_status": by_status,
        "total_gross": round(tot_gross, 2),
        "total_bonus": round(tot_bonus, 2),
        "total_deduction": round(tot_ded, 2),
        "total_net": round(tot_net, 2),
        "per_driver": driver_rows,
        "sdm_vs_revenue": {"payroll_cost": payroll_cost, "revenue": revenue, "ratio_pct": ratio},
    }


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


# --------------------------------------------------------------------------- PDF
def report_pdf(data: dict, company: dict = None) -> bytes:
    company = company or {}
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm, title="Rekap Payroll")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=17, spaceAfter=2)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#6B6B73"))
    el = [Paragraph(f"{company.get('name') or 'Rekap Payroll Driver'}", h1),
          Paragraph(f"Periode {data.get('period')} · SDM vs Revenue: {_rp(data['sdm_vs_revenue']['payroll_cost'])} / {_rp(data['sdm_vs_revenue']['revenue'])} = {data['sdm_vs_revenue']['ratio_pct']}%", small),
          Spacer(1, 8)]

    bs = data.get("by_status", {})
    summ = [["Total Payout", "Draft", "Disetujui", "Dibayar", "Total Gross", "Total Bonus", "Total Potongan", "Total Net"],
            [str(data.get("count", 0)), str(bs.get("draft", 0)), str(bs.get("approved", 0)), str(bs.get("paid", 0)),
             _rp(data.get("total_gross")), _rp(data.get("total_bonus")), _rp(data.get("total_deduction")), _rp(data.get("total_net"))]]
    ts = Table(summ, repeatRows=1)
    ts.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")), ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    el.append(ts)
    el.append(Spacer(1, 12))

    rows = [["Driver", "Trip", "KM", "Gross", "Bonus", "Potongan", "Total", "YTD", "Status"]]
    for r in data.get("per_driver", []):
        rows.append([r["driver_name"], str(r["trips"]), str(r["km"]), _rp(r["gross"]), _rp(r["bonus"]),
                     _rp(r["deduction"]), _rp(r["total"]), _rp(r["ytd"]), ", ".join(r["statuses"])])
    if len(rows) == 1:
        rows.append(["(belum ada payout pada periode ini)", "", "", "", "", "", "", "", ""])
    dt = Table(rows, repeatRows=1)
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058CC")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")),
        ("ALIGN", (1, 0), (7, -1), "RIGHT"), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFB")])]))
    el.append(dt)
    doc.build(el)
    return buf.getvalue()


# --------------------------------------------------------------------------- Excel
def report_xlsx(data: dict, company: dict = None) -> bytes:
    wb = Workbook()
    hfill = PatternFill("solid", fgColor="1C1C1E")
    hfont = Font(color="FFFFFF", bold=True)

    def head(ws, n):
        for c in range(1, n + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 40)

    ws = wb.active
    ws.title = "Ringkasan"
    bs = data.get("by_status", {})
    sr = data.get("sdm_vs_revenue", {})
    ws.append(["Metrik", "Nilai"]); head(ws, 2)
    for k, v in [("Periode", data.get("period")), ("Total Payout", data.get("count")),
                 ("Draft", bs.get("draft", 0)), ("Disetujui", bs.get("approved", 0)), ("Dibayar", bs.get("paid", 0)),
                 ("Total Gross", data.get("total_gross")), ("Total Bonus", data.get("total_bonus")),
                 ("Total Potongan", data.get("total_deduction")), ("Total Net", data.get("total_net")),
                 ("Biaya SDM (periode)", sr.get("payroll_cost")), ("Revenue (periode)", sr.get("revenue")),
                 ("Rasio SDM/Revenue (%)", sr.get("ratio_pct"))]:
        ws.append([k, v])

    ws2 = wb.create_sheet("Per Driver")
    ws2.append(["Driver", "Trip", "KM", "Gross", "Bonus", "Potongan", "Total", "YTD", "Status"]); head(ws2, 9)
    for r in data.get("per_driver", []):
        ws2.append([r["driver_name"], r["trips"], r["km"], r["gross"], r["bonus"], r["deduction"],
                    r["total"], r["ytd"], ", ".join(r["statuses"])])
    autosize(ws); autosize(ws2)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
