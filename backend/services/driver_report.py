"""services/driver_report.py — Laporan Driver (Leaderboard) + Revenue-per-Trip (E13).

Agregasi read-only dari koleksi `trips` (+ nama dari `drivers`):
  - Leaderboard per-driver pada periode: total trip, trip selesai, %selesai, total KM,
    total revenue, revenue-per-trip, rata-rata KM per trip → diperingkat by revenue.
  - Ringkasan armada (fleet): total trip/selesai/km/revenue + revenue-per-trip +
    rata-rata KM per trip + jumlah driver aktif.

Periode = bulan efektif trip (end_at → start_at → created_at, ambil YYYY-MM).
Tidak menulis DB. Export PDF (reportlab) + Excel (openpyxl).
"""
from datetime import datetime, timezone
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ekspresi bulan efektif trip: end_at → start_at → created_at (YYYY-MM)
_EFF_MONTH = {
    "$substrBytes": [
        {"$ifNull": ["$end_at", {"$ifNull": ["$start_at", {"$ifNull": ["$created_at", ""]}]}]},
        0, 7,
    ]
}


async def build(db, period=None) -> dict:
    period = period or datetime.now(timezone.utc).strftime("%Y-%m")

    pipeline = [
        {"$addFields": {"_eff": _EFF_MONTH}},
        {"$match": {"_eff": period, "driver_id": {"$ne": None}}},
        {"$group": {
            "_id": "$driver_id",
            "trips": {"$sum": 1},
            "completed": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}},
            "km": {"$sum": {"$ifNull": ["$distance_km", 0]}},
            "revenue": {"$sum": {"$ifNull": ["$revenue", 0]}},
            "completed_revenue": {"$sum": {
                "$cond": [{"$eq": ["$status", "completed"]}, {"$ifNull": ["$revenue", 0]}, 0]}},
        }},
        {"$sort": {"revenue": -1, "trips": -1}},
    ]
    rows = await db.trips.aggregate(pipeline).to_list(2000)

    # Peta nama driver (bounded).
    name_map = {
        d["id"]: d.get("name")
        for d in await db.drivers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(3000)
    }

    drivers = []
    tot_trips = tot_completed = 0
    tot_km = tot_rev = 0.0
    for i, r in enumerate(rows):
        trips = int(r.get("trips") or 0)
        completed = int(r.get("completed") or 0)
        km = float(r.get("km") or 0)
        rev = float(r.get("revenue") or 0)
        drivers.append({
            "rank": i + 1,
            "driver_id": r["_id"],
            "driver_name": name_map.get(r["_id"]) or r["_id"],
            "trips": trips,
            "completed": completed,
            "completion_rate": round(completed / trips * 100, 1) if trips else 0.0,
            "km": round(km, 1),
            "revenue": round(rev, 2),
            "revenue_per_trip": round(rev / trips, 2) if trips else 0.0,
            "avg_km_per_trip": round(km / trips, 1) if trips else 0.0,
        })
        tot_trips += trips
        tot_completed += completed
        tot_km += km
        tot_rev += rev

    fleet = {
        "trips": tot_trips,
        "completed": tot_completed,
        "completion_rate": round(tot_completed / tot_trips * 100, 1) if tot_trips else 0.0,
        "km": round(tot_km, 1),
        "revenue": round(tot_rev, 2),
        "revenue_per_trip": round(tot_rev / tot_trips, 2) if tot_trips else 0.0,
        "avg_km_per_trip": round(tot_km / tot_trips, 1) if tot_trips else 0.0,
        "active_drivers": len(drivers),
    }

    return {"period": period, "fleet": fleet, "drivers": drivers}


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


# --------------------------------------------------------------------------- PDF
def report_pdf(data: dict, company: dict = None) -> bytes:
    company = company or {}
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm, title="Laporan Driver")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=17, spaceAfter=2)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9.5,
                           textColor=colors.HexColor("#6B6B73"))
    fl = data.get("fleet", {})
    el = [Paragraph(f"{company.get('name') or 'Laporan Driver (Leaderboard)'}", h1),
          Paragraph(f"Periode {data.get('period')} · Total {fl.get('trips', 0)} trip "
                    f"({fl.get('completed', 0)} selesai) · Revenue {_rp(fl.get('revenue'))} · "
                    f"Revenue/Trip {_rp(fl.get('revenue_per_trip'))}", small),
          Spacer(1, 10)]

    rows = [["#", "Driver", "Trip", "Selesai", "% Selesai", "Total KM",
             "Revenue", "Revenue/Trip", "KM/Trip"]]
    for r in data.get("drivers", []):
        rows.append([str(r["rank"]), r["driver_name"], str(r["trips"]), str(r["completed"]),
                     f"{r['completion_rate']}%", str(r["km"]), _rp(r["revenue"]),
                     _rp(r["revenue_per_trip"]), str(r["avg_km_per_trip"])])
    if len(rows) == 1:
        rows.append(["", "(belum ada trip pada periode ini)", "", "", "", "", "", "", ""])
    dt = Table(rows, repeatRows=1)
    dt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058CC")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"), ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FAFAFB")])]))
    el.append(dt)
    doc.build(el)
    return buf.getvalue()


# --------------------------------------------------------------------------- Excel
def report_xlsx(data: dict, company: dict = None) -> bytes:
    wb = Workbook()
    hfill = PatternFill("solid", fgColor="0058CC")
    hfont = Font(color="FFFFFF", bold=True)

    def head(ws, n):
        for c in range(1, n + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 40)

    fl = data.get("fleet", {})
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Metrik", "Nilai"]); head(ws, 2)
    for k, v in [("Periode", data.get("period")), ("Total Trip", fl.get("trips", 0)),
                 ("Trip Selesai", fl.get("completed", 0)), ("% Selesai", fl.get("completion_rate", 0)),
                 ("Total KM", fl.get("km", 0)), ("Total Revenue", fl.get("revenue", 0)),
                 ("Revenue per Trip", fl.get("revenue_per_trip", 0)),
                 ("Rata-rata KM per Trip", fl.get("avg_km_per_trip", 0)),
                 ("Driver Aktif", fl.get("active_drivers", 0))]:
        ws.append([k, v])

    ws2 = wb.create_sheet("Leaderboard Driver")
    ws2.append(["#", "Driver", "Trip", "Selesai", "% Selesai", "Total KM",
                "Revenue", "Revenue/Trip", "KM/Trip"]); head(ws2, 9)
    for r in data.get("drivers", []):
        ws2.append([r["rank"], r["driver_name"], r["trips"], r["completed"], r["completion_rate"],
                    r["km"], r["revenue"], r["revenue_per_trip"], r["avg_km_per_trip"]])
    autosize(ws); autosize(ws2)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
