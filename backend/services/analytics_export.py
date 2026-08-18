"""services/analytics_export.py — E4 cockpit export (PDF reportlab + Excel openpyxl).

Read-only: menerima dict hasil analytics.assemble_cockpit(db, rng) lalu merangkainya
menjadi dokumen ringkasan eksekutif. Tidak menyentuh DB.
"""
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


# --------------------------------------------------------------------------- PDF
def cockpit_pdf(data: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm, title="BI Cockpit")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=18, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#1C1C1E"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B6B73"))
    el = []
    rng = data.get("range", {})
    el.append(Paragraph("BI &amp; Management Cockpit", h1))
    el.append(Paragraph(f"Periode: {rng.get('label', '-')}", small))
    el.append(Spacer(1, 8))

    # KPI
    m = (data.get("summary") or {}).get("metrics", {})

    def mv(k):
        return (m.get(k) or {}).get("value", 0)
    el.append(Paragraph("Ringkasan Eksekutif", h2))
    kpi_rows = [
        ["Pendapatan", _rp(mv("revenue")), "Pengeluaran", _rp(mv("expenses"))],
        ["Laba", _rp(mv("profit")), "Margin", f"{mv('margin')}%"],
        ["Booking", str(mv("bookings")), "Lead", str(mv("leads"))],
        ["Konversi", f"{mv('conversion_rate')}%", "Piutang (AR)", _rp(mv("outstanding_ar"))],
    ]
    t = Table(kpi_rows, colWidths=[35 * mm, 45 * mm, 35 * mm, 45 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6B6B73")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#6B6B73")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#EFF0F2")),
    ]))
    el.append(t)
    el.append(Spacer(1, 10))

    # Funnel
    el.append(Paragraph("Sales Funnel", h2))
    frows = [["Tahap", "Jumlah", "Rate"]]
    for s in (data.get("funnel") or {}).get("stages", []):
        frows.append([s.get("label"), str(s.get("count")), f"{s.get('rate')}%"])
    el.append(_grid(frows, [70 * mm, 40 * mm, 40 * mm]))
    el.append(Spacer(1, 10))

    # Channels
    el.append(Paragraph("Channel Mix &amp; ROAS", h2))
    crows = [["Channel", "Lead", "Menang", "Spend", "CPL", "ROAS"]]
    for c in (data.get("channels") or {}).get("channels", []):
        crows.append([c.get("channel"), str(c.get("leads")), str(c.get("won")),
                      _rp(c.get("spend")), _rp(c.get("cpl")) if c.get("cpl") else "-",
                      f"{c.get('roas')}x" if c.get("roas") else "-"])
    el.append(_grid(crows, [38 * mm, 18 * mm, 22 * mm, 30 * mm, 30 * mm, 22 * mm]))
    el.append(Spacer(1, 10))

    # Fleet
    el.append(Paragraph("Fleet ROI", h2))
    vrows = [["Armada", "Trip", "Pendapatan", "Biaya", "Laba", "ROI"]]
    for v in (data.get("fleet") or {}).get("vehicles", [])[:12]:
        vrows.append([v.get("vehicle_name"), str(v.get("trips")), _rp(v.get("revenue")),
                      _rp(v.get("expenses")), _rp(v.get("profit")),
                      f"{v.get('roi_pct')}%" if v.get("roi_pct") is not None else "-"])
    el.append(_grid(vrows, [42 * mm, 16 * mm, 30 * mm, 26 * mm, 26 * mm, 20 * mm]))
    el.append(Spacer(1, 10))

    # AR aging
    el.append(Paragraph("AR Aging", h2))
    arows = [["Bucket", "Jumlah", "Nilai"]]
    for b in (data.get("ar_aging") or {}).get("buckets", []):
        arows.append([b.get("label"), str(b.get("count")), _rp(b.get("amount"))])
    el.append(_grid(arows, [70 * mm, 40 * mm, 40 * mm]))

    doc.build(el)
    return buf.getvalue()


def _grid(rows, widths):
    t = Table(rows, colWidths=widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F1F4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")),
    ]))
    return t


# --------------------------------------------------------------------------- Excel
def cockpit_xlsx(data: dict) -> bytes:
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1C1C1E")
    head_font = Font(color="FFFFFF", bold=True)

    def style_head(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 40)

    rng = data.get("range", {})
    m = (data.get("summary") or {}).get("metrics", {})
    ws = wb.active
    ws.title = "Ringkasan"
    ws.append(["Metrik", "Nilai", "Periode Lalu", "Delta %"])
    style_head(ws, 4)
    label_map = {"revenue": "Pendapatan", "expenses": "Pengeluaran", "profit": "Laba",
                 "margin": "Margin %", "bookings": "Booking", "leads": "Lead",
                 "won": "Deal Menang", "conversion_rate": "Konversi %",
                 "avg_booking_value": "Rata-rata Booking", "outstanding_ar": "Piutang (AR)",
                 "ar_count": "Jumlah AR"}
    for k, lab in label_map.items():
        if k in m:
            ws.append([lab, m[k].get("value"), m[k].get("prev"), m[k].get("delta_pct")])
    ws.append([])
    ws.append(["Periode", rng.get("label", "")])
    autosize(ws)

    ws2 = wb.create_sheet("Funnel")
    ws2.append(["Tahap", "Jumlah", "Rate %"])
    style_head(ws2, 3)
    for s in (data.get("funnel") or {}).get("stages", []):
        ws2.append([s.get("label"), s.get("count"), s.get("rate")])
    autosize(ws2)

    ws3 = wb.create_sheet("Channel ROI")
    ws3.append(["Channel", "Lead", "Menang", "Pendapatan", "Spend", "CPL", "CAC", "ROAS"])
    style_head(ws3, 8)
    for c in (data.get("channels") or {}).get("channels", []):
        ws3.append([c.get("channel"), c.get("leads"), c.get("won"), c.get("revenue"),
                    c.get("spend"), c.get("cpl"), c.get("cac"), c.get("roas")])
    autosize(ws3)

    ws4 = wb.create_sheet("Fleet ROI")
    ws4.append(["Armada", "Trip", "Jarak (km)", "Pendapatan", "Biaya", "Laba", "ROI %"])
    style_head(ws4, 7)
    for v in (data.get("fleet") or {}).get("vehicles", []):
        ws4.append([v.get("vehicle_name"), v.get("trips"), v.get("distance_km"),
                    v.get("revenue"), v.get("expenses"), v.get("profit"), v.get("roi_pct")])
    autosize(ws4)

    ws5 = wb.create_sheet("Driver")
    ws5.append(["Driver", "Trip", "Selesai", "Completion %", "Pendapatan", "Jarak (km)"])
    style_head(ws5, 6)
    for d in (data.get("drivers") or {}).get("drivers", []):
        ws5.append([d.get("driver_name"), d.get("trips"), d.get("completed"),
                    d.get("completion_rate"), d.get("revenue"), d.get("distance_km")])
    autosize(ws5)

    ws6 = wb.create_sheet("AR Aging")
    ws6.append(["Bucket", "Jumlah", "Nilai"])
    style_head(ws6, 3)
    for b in (data.get("ar_aging") or {}).get("buckets", []):
        ws6.append([b.get("label"), b.get("count"), b.get("amount")])
    autosize(ws6)

    ws7 = wb.create_sheet("Forecast")
    ws7.append(["Bulan", "Tipe", "Nilai"])
    style_head(ws7, 3)
    fc = data.get("forecast") or {}
    for h in fc.get("history", []):
        ws7.append([h.get("month"), "aktual", h.get("value")])
    for f in fc.get("forecast", []):
        ws7.append([f.get("month"), "prediksi", f.get("value")])
    autosize(ws7)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
