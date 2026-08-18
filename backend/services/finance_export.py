"""services/finance_export.py — E5 laporan keuangan (PDF reportlab + Excel openpyxl).

Menerima dict dari finance_automation.assemble_finance_report(db, period):
{period, pl, ar_aging, cashflow, reconciliation}. Tidak menyentuh DB.
"""
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

CAT_LABEL = {"bbm": "BBM", "tol": "Tol", "uang_jalan": "Uang Jalan", "other": "Lainnya"}
MONTHS_ID = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


def _mlabel(key):
    try:
        return f"{MONTHS_ID[int(str(key)[5:7]) - 1]} {str(key)[:4]}"
    except Exception:
        return str(key)


def _grid(rows, widths, aligns=None):
    t = Table(rows, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4), ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#EFF0F2")),
    ]
    for col in (aligns or []):
        style.append(("ALIGN", (col, 1), (col, -1), "RIGHT"))
    t.setStyle(TableStyle(style))
    return t


# --------------------------------------------------------------------------- PDF
def finance_report_pdf(data: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm, title="Laporan Keuangan")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=18, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B6B73"))
    el = []
    pl = data.get("pl", {})
    el.append(Paragraph("Laporan Keuangan", h1))
    el.append(Paragraph(f"Periode: {_mlabel(data.get('period'))}", small))
    el.append(Spacer(1, 8))

    el.append(Paragraph("Laba-Rugi (incl. Perawatan)", h2))
    pl_rows = [
        ["Pendapatan (kas masuk)", _rp(pl.get("revenue"))],
        ["Biaya Operasional", _rp(pl.get("operational_expenses"))],
        ["Biaya Perawatan", _rp(pl.get("maintenance_cost"))],
        ["Total Biaya", _rp(pl.get("total_cost"))],
        ["Laba Bersih", _rp(pl.get("profit"))],
        ["Margin", f"{pl.get('margin', 0)}%"],
    ]
    el.append(_grid(pl_rows, [90 * mm, 70 * mm], aligns=[1]))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Laba-Rugi per Unit", h2))
    u_rows = [["Armada", "Pendapatan", "Operasional", "Perawatan", "Laba"]]
    for u in pl.get("per_unit", [])[:14]:
        u_rows.append([u.get("vehicle_name"), _rp(u.get("revenue")), _rp(u.get("operational_expenses")),
                       _rp(u.get("maintenance")), _rp(u.get("profit"))])
    el.append(_grid(u_rows, [44 * mm, 30 * mm, 30 * mm, 28 * mm, 28 * mm], aligns=[1, 2, 3, 4]))
    el.append(Spacer(1, 10))

    el.append(Paragraph("AR Aging (Piutang)", h2))
    a_rows = [["Bucket", "Jumlah", "Nilai"]]
    for b in data.get("ar_aging", {}).get("buckets", []):
        a_rows.append([b.get("label"), str(b.get("count")), _rp(b.get("amount"))])
    el.append(_grid(a_rows, [70 * mm, 40 * mm, 40 * mm], aligns=[1, 2]))
    el.append(Spacer(1, 10))

    el.append(Paragraph("Arus Kas", h2))
    c_rows = [["Bulan", "Kas Masuk", "Kas Keluar", "Net", "Saldo"]]
    for m in data.get("cashflow", {}).get("months", []):
        c_rows.append([_mlabel(m.get("month")), _rp(m.get("cash_in")), _rp(m.get("cash_out")),
                       _rp(m.get("net")), _rp(m.get("balance"))])
    for p in data.get("cashflow", {}).get("projection", []):
        c_rows.append([f"{_mlabel(p.get('month'))} (proy)", "-", "-", _rp(p.get("net")), _rp(p.get("balance"))])
    el.append(_grid(c_rows, [38 * mm, 32 * mm, 32 * mm, 28 * mm, 30 * mm], aligns=[1, 2, 3, 4]))

    doc.build(el)
    return buf.getvalue()


# --------------------------------------------------------------------------- Excel
def finance_report_xlsx(data: dict) -> bytes:
    wb = Workbook()
    hfill = PatternFill("solid", fgColor="1C1C1E")
    hfont = Font(color="FFFFFF", bold=True)

    def head(ws, n):
        for c in range(1, n + 1):
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).font = hfont
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 38)

    pl = data.get("pl", {})
    ws = wb.active
    ws.title = "Laba-Rugi"
    ws.append(["Pos", "Nilai"])
    head(ws, 2)
    for lab, key in [("Pendapatan", "revenue"), ("Biaya Operasional", "operational_expenses"),
                     ("Biaya Perawatan", "maintenance_cost"), ("Total Biaya", "total_cost"),
                     ("Laba Bersih", "profit"), ("Margin %", "margin")]:
        ws.append([lab, pl.get(key)])
    ws.append([])
    ws.append(["Periode", data.get("period")])
    autosize(ws)

    ws2 = wb.create_sheet("Per Unit")
    ws2.append(["Armada", "Pendapatan", "Operasional", "Perawatan", "Laba"])
    head(ws2, 5)
    for u in pl.get("per_unit", []):
        ws2.append([u.get("vehicle_name"), u.get("revenue"), u.get("operational_expenses"),
                    u.get("maintenance"), u.get("profit")])
    autosize(ws2)

    ws3 = wb.create_sheet("AR Aging")
    ws3.append(["Bucket", "Jumlah", "Nilai"])
    head(ws3, 3)
    for b in data.get("ar_aging", {}).get("buckets", []):
        ws3.append([b.get("label"), b.get("count"), b.get("amount")])
    autosize(ws3)

    ws4 = wb.create_sheet("Arus Kas")
    ws4.append(["Bulan", "Tipe", "Kas Masuk", "Kas Keluar", "Net", "Saldo"])
    head(ws4, 6)
    for m in data.get("cashflow", {}).get("months", []):
        ws4.append([m.get("month"), "aktual", m.get("cash_in"), m.get("cash_out"), m.get("net"), m.get("balance")])
    for p in data.get("cashflow", {}).get("projection", []):
        ws4.append([p.get("month"), "proyeksi", None, None, p.get("net"), p.get("balance")])
    autosize(ws4)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
