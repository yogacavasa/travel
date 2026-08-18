"""services/exporter.py — generator file export (Phase 5).

- invoice_pdf(inv)  : faktur PDF (reportlab).
- invoice_xlsx(inv) : faktur Excel (openpyxl).
- report_xlsx(data) : ringkasan laporan Excel (openpyxl).
Semua mengembalikan bytes (dipakai StreamingResponse).
"""
from datetime import datetime, timezone
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

BRAND = colors.HexColor("#007AFF")
INK = colors.HexColor("#1C1C1E")
MUTED = colors.HexColor("#6B6B73")
LINE = colors.HexColor("#EFF0F2")
COMPANY = "RahazaTrans"

STATUS_LABEL = {"draft": "Draft", "sent": "Terkirim", "paid": "Lunas"}
QUO_STATUS_LABEL = {"draft": "Draft", "sent": "Terkirim", "accepted": "Diterima",
                    "rejected": "Ditolak", "expired": "Kedaluwarsa", "converted": "Jadi Booking"}


def _rp(v) -> str:
    return "Rp " + f"{int(round(float(v or 0))):,}".replace(",", ".")


def _fdate(value) -> str:
    if not value:
        return "-"
    try:
        d = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return d.strftime("%d %b %Y")
    except Exception:
        return str(value)[:10]


def _esc(v) -> str:
    """EXPORT-1 fix: escape teks user sebelum masuk markup reportlab Paragraph.
    Tanpa ini, karakter '<' '&' '>' (mis. nama/catatan/destinasi berisi markup) membuat
    parser mini-HTML reportlab melempar → HTTP 500. Escape → aman & tampil apa adanya."""
    from xml.sax.saxutils import escape as _x
    return _x("" if v is None else str(v))


def invoice_pdf(inv: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm, title=str(inv.get("number", "INVOICE")),
    )
    ss = getSampleStyleSheet()
    h_company = ParagraphStyle("co", parent=ss["Title"], fontSize=20, textColor=BRAND, spaceAfter=2)
    h_sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, textColor=MUTED)
    h_label = ParagraphStyle("lbl", parent=ss["Normal"], fontSize=8, textColor=MUTED, spaceAfter=1)
    h_val = ParagraphStyle("val", parent=ss["Normal"], fontSize=11, textColor=INK, spaceAfter=6)
    h_invtitle = ParagraphStyle("it", parent=ss["Title"], fontSize=16, textColor=INK, alignment=2)

    story = []
    head = Table(
        [[Paragraph(COMPANY, h_company), Paragraph("INVOICE", h_invtitle)],
         [Paragraph("Fleet &amp; Travel Management", h_sub),
          Paragraph(f"<b>{_esc(inv.get('number','-'))}</b>", ParagraphStyle('n', parent=h_sub, alignment=2, fontSize=10))]],
        colWidths=[95 * mm, 75 * mm],
    )
    head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(head)
    story.append(Spacer(1, 10))
    story.append(Table([[""]], colWidths=[170 * mm],
                       style=TableStyle([("LINEBELOW", (0, 0), (-1, -1), 1, LINE)])))
    story.append(Spacer(1, 12))

    meta = Table(
        [[Paragraph("DITAGIHKAN KEPADA", h_label), Paragraph("STATUS", h_label)],
         [Paragraph(f"<b>{_esc(inv.get('customer_name','-'))}</b>", h_val),
          Paragraph(STATUS_LABEL.get(inv.get("status"), inv.get("status", "-")), h_val)],
         [Paragraph("BOOKING", h_label), Paragraph("TANGGAL TERBIT", h_label)],
         [Paragraph(_esc(inv.get("booking_code", "-") or "-"), h_val), Paragraph(_fdate(inv.get("issued_at")), h_val)],
         [Paragraph("JATUH TEMPO", h_label), Paragraph("", h_label)],
         [Paragraph(_fdate(inv.get("due_at")), h_val), Paragraph("", h_val)]],
        colWidths=[85 * mm, 85 * mm],
    )
    meta.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
    story.append(meta)
    story.append(Spacer(1, 14))

    amount = float(inv.get("amount", 0) or 0)
    desc = f"Sewa armada — {inv.get('booking_code','')}".strip()
    items = Table(
        [["Deskripsi", "Jumlah"],
         [desc or "Layanan sewa armada", _rp(amount)],
         ["", ""],
         ["Total", _rp(amount)]],
        colWidths=[120 * mm, 50 * mm],
    )
    items.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEBELOW", (0, 1), (-1, 1), 0.5, LINE),
        ("LINEABOVE", (0, 3), (-1, 3), 1, INK),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, -1), INK),
    ]))
    story.append(items)
    story.append(Spacer(1, 18))
    if inv.get("notes"):
        story.append(Paragraph(f"<font color='#6B6B73' size='9'>Catatan: {_esc(inv.get('notes'))}</font>", h_sub))
        story.append(Spacer(1, 8))
    story.append(Paragraph(
        "<font color='#8E8E93' size='8'>Dokumen ini dibuat otomatis oleh sistem RahazaTrans.</font>", h_sub))

    doc.build(story)
    return buf.getvalue()


def quotation_pdf(quo: dict) -> bytes:
    """Penawaran (QUO) PDF — ber-item, gaya konsisten dgn invoice (B2)."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm, title=str(quo.get("number", "PENAWARAN")),
    )
    ss = getSampleStyleSheet()
    h_company = ParagraphStyle("co", parent=ss["Title"], fontSize=20, textColor=BRAND, spaceAfter=2)
    h_sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, textColor=MUTED)
    h_label = ParagraphStyle("lbl", parent=ss["Normal"], fontSize=8, textColor=MUTED, spaceAfter=1)
    h_val = ParagraphStyle("val", parent=ss["Normal"], fontSize=11, textColor=INK, spaceAfter=6)
    h_title = ParagraphStyle("it", parent=ss["Title"], fontSize=16, textColor=INK, alignment=2)

    story = []
    head = Table(
        [[Paragraph(COMPANY, h_company), Paragraph("PENAWARAN", h_title)],
         [Paragraph("Fleet &amp; Travel Management", h_sub),
          Paragraph(f"<b>{quo.get('number','-')}</b>", ParagraphStyle('n', parent=h_sub, alignment=2, fontSize=10))]],
        colWidths=[95 * mm, 75 * mm],
    )
    head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(head)
    story.append(Spacer(1, 10))
    story.append(Table([[""]], colWidths=[170 * mm],
                       style=TableStyle([("LINEBELOW", (0, 0), (-1, -1), 1, LINE)])))
    story.append(Spacer(1, 12))

    meta = Table(
        [[Paragraph("UNTUK", h_label), Paragraph("STATUS", h_label)],
         [Paragraph(f"<b>{_esc(quo.get('customer_name','-'))}</b>", h_val),
          Paragraph(QUO_STATUS_LABEL.get(quo.get("status"), quo.get("status", "-")), h_val)],
         [Paragraph("DESTINASI", h_label), Paragraph("TANGGAL TRIP", h_label)],
         [Paragraph(_esc(quo.get("destination", "-") or "-"), h_val), Paragraph(_fdate(quo.get("trip_date")), h_val)],
         [Paragraph("BERLAKU HINGGA", h_label), Paragraph("PAX", h_label)],
         [Paragraph(_fdate(quo.get("valid_until")), h_val), Paragraph(str(quo.get("pax", "-") or "-"), h_val)]],
        colWidths=[85 * mm, 85 * mm],
    )
    meta.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOTTOMPADDING", (0, 0), (-1, -1), 2)]))
    story.append(meta)
    story.append(Spacer(1, 14))

    rows = [["Deskripsi", "Jumlah"]]
    for it in (quo.get("items") or []):
        rows.append([str(it.get("label", "")), _rp(it.get("amount"))])
    rows.append(["Total", _rp(quo.get("total", 0))])
    items = Table(rows, colWidths=[120 * mm, 50 * mm])
    last = len(rows) - 1
    items.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LINEABOVE", (0, last), (-1, last), 1, INK),
        ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, -1), INK),
    ]))
    story.append(items)
    story.append(Spacer(1, 18))
    if quo.get("notes"):
        story.append(Paragraph(f"<font color='#6B6B73' size='9'>Catatan: {_esc(quo.get('notes'))}</font>", h_sub))
        story.append(Spacer(1, 8))
    story.append(Paragraph(
        "<font color='#8E8E93' size='8'>Penawaran dibuat otomatis oleh sistem RahazaTrans. "
        "Harga dapat berubah sesuai ketersediaan armada.</font>", h_sub))

    doc.build(story)
    return buf.getvalue()


def _autosize(ws):
    for col in ws.columns:
        width = 10
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                width = max(width, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(width, 40)


def invoice_xlsx(inv: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    bold = Font(bold=True)
    title = Font(bold=True, size=14, color="007AFF")
    ws["A1"] = COMPANY
    ws["A1"].font = title
    ws["A2"] = "INVOICE"
    ws["A2"].font = bold
    rows = [
        ("Nomor", inv.get("number", "-")),
        ("Customer", inv.get("customer_name", "-")),
        ("Booking", inv.get("booking_code", "-")),
        ("Status", STATUS_LABEL.get(inv.get("status"), inv.get("status", "-"))),
        ("Tanggal Terbit", _fdate(inv.get("issued_at"))),
        ("Jatuh Tempo", _fdate(inv.get("due_at"))),
        ("Jumlah", _rp(inv.get("amount"))),
    ]
    r = 4
    for label, val in rows:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=str(val))
        r += 1
    _autosize(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_xlsx(data: dict) -> bytes:
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="007AFF")
    head_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    def _header(ws, cols):
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center")

    ws = wb.active
    ws.title = "Ringkasan"
    ws["A1"] = f"Laporan Keuangan — {COMPANY}"
    ws["A1"].font = Font(bold=True, size=13, color="007AFF")
    ws["A2"] = f"Periode: {data.get('period','-')}"
    summ = [
        ("Total Pemasukan", _rp(data.get("revenue_total"))),
        ("Total Pengeluaran", _rp(data.get("expense_total"))),
        ("Laba Bersih", _rp(data.get("profit_total"))),
        ("Piutang (AR)", _rp(data.get("outstanding_ar"))),
    ]
    r = 4
    for label, val in summ:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1
    _autosize(ws)

    ws2 = wb.create_sheet("Tren Bulanan")
    _header(ws2, ["Bulan", "Pemasukan", "Pengeluaran", "Laba"])
    for row in data.get("revenue_by_month", []):
        ws2.append([row.get("month"), float(row.get("revenue", 0)),
                    float(row.get("expenses", 0)), float(row.get("profit", 0))])
    _autosize(ws2)

    ws3 = wb.create_sheet("Kategori Pengeluaran")
    _header(ws3, ["Kategori", "Jumlah"])
    for row in data.get("expense_by_category", []):
        ws3.append([row.get("category"), float(row.get("amount", 0))])
    _autosize(ws3)

    ws4 = wb.create_sheet("Top Customer")
    _header(ws4, ["Customer", "Nilai (Rp)", "Total Trip"])
    for row in data.get("top_customers", []):
        ws4.append([row.get("name"), float(row.get("lifetime_value", 0)), int(row.get("total_trips", 0))])
    _autosize(ws4)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
