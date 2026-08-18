"""services/calendar_export.py — Ekspor Kalender Keberangkatan (PDF reportlab + Excel openpyxl).

Read-only: menerima label periode + daftar booking (dict), merangkainya menjadi
lembar jadwal keberangkatan yang siap cetak / dibagikan. Mendukung pengelompokan
per armada (satu bagian / sheet per kendaraan). Tidak menyentuh DB.
"""
import re
from datetime import datetime
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "Mei", "Jun", "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]
_MONTHS_FULL = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
STATUS_ID = {"pending": "Menunggu", "hold": "Hold", "confirmed": "Terkonfirmasi", "ongoing": "Berjalan", "completed": "Selesai", "cancelled": "Dibatalkan", "draft": "Draft"}
PAY_ID = {"lunas": "Lunas", "dp": "DP", "belum_bayar": "Belum Bayar", "selesai": "Selesai"}
COLS = ["Kode", "Berangkat", "Kembali", "Pelanggan", "Armada", "Driver", "Rute", "Status", "Bayar", "Total"]
COLS_GROUPED = ["Kode", "Berangkat", "Kembali", "Pelanggan", "Driver", "Rute", "Status", "Bayar", "Total"]  # tanpa Armada


def _rp(v):
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except Exception:
        return "Rp 0"


def _fmt_dt(iso):
    if not iso:
        return "-"
    try:
        dt = datetime.fromisoformat(str(iso).replace("Z", "+00:00"))
        return f"{dt.day:02d} {_MONTHS[dt.month - 1]} {dt.year} {dt.hour:02d}:{dt.minute:02d}"
    except Exception:
        return str(iso)[:16]


def _fmt_date(d):
    try:
        y, m, dd = str(d)[:10].split("-")
        return f"{int(dd):02d} {_MONTHS[int(m) - 1]} {y}"
    except Exception:
        return str(d)


def _month_title(label):
    try:
        y, m = str(label).split("-")
        return f"{_MONTHS_FULL[int(m) - 1]} {y}"
    except Exception:
        return str(label)


def period_label(month=None, start=None, end=None):
    if start and end:
        return f"{_fmt_date(start)} \u2013 {_fmt_date(end)}"
    if month:
        return _month_title(month)
    return "Semua Periode"


def _route(b):
    o = (b.get("origin") or "").strip()
    d = (b.get("destination") or "").strip()
    if o and d:
        return f"{o} \u2192 {d}"
    return d or o or "-"


def _values(b, include_vehicle=True):
    base = [
        b.get("code") or "-",
        _fmt_dt(b.get("start_datetime")),
        _fmt_dt(b.get("end_datetime")),
        b.get("customer_name") or "-",
    ]
    if include_vehicle:
        base.append(b.get("vehicle_name") or "-")
    base += [
        b.get("driver_name") or "-",
        _route(b),
        STATUS_ID.get(b.get("status"), b.get("status") or "-"),
        PAY_ID.get(b.get("payment_status"), b.get("payment_status") or "-"),
        _rp(b.get("total_amount")),
    ]
    return base


def _group_rows(rows):
    """Kelompokkan booking per armada; kembalikan list (nama_armada, [rows]) terurut."""
    groups = {}
    for b in rows:
        key = b.get("vehicle_name") or "Tanpa Armada"
        groups.setdefault(key, []).append(b)
    return sorted(groups.items(), key=lambda kv: kv[0].lower())


# --------------------------------------------------------------------------- PDF
def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", parent=styles["Title"], fontSize=16, spaceAfter=2),
        "sub": ParagraphStyle("sub", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B6B73")),
        "grp": ParagraphStyle("grp", parent=styles["Heading2"], fontSize=11, spaceBefore=8, spaceAfter=3, textColor=colors.HexColor("#0058CC")),
        "cell": ParagraphStyle("cell", parent=styles["Normal"], fontSize=7.2, leading=8.6),
        "head": ParagraphStyle("head", parent=styles["Normal"], fontSize=7.4, leading=9, textColor=colors.white, fontName="Helvetica-Bold"),
    }


def _pdf_table(rows, st, include_vehicle=True):
    cols = COLS if include_vehicle else COLS_GROUPED
    data = [[Paragraph(c, st["head"]) for c in cols]]
    for b in rows:
        data.append([Paragraph(str(v), st["cell"]) for v in _values(b, include_vehicle)])
    if len(data) == 1:
        data.append([Paragraph("Tidak ada keberangkatan.", st["cell"])] + [Paragraph("", st["cell"]) for _ in range(len(cols) - 1)])
    widths = [16, 30, 30, 34, 28, 26, 40, 26, 20, 27] if include_vehicle else [18, 32, 32, 40, 30, 46, 28, 22, 29]
    tbl = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0058CC")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8DAE0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return tbl


def calendar_pdf(period: str, rows: list, grouped: bool = False) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4), topMargin=12 * mm, bottomMargin=12 * mm,
        leftMargin=10 * mm, rightMargin=10 * mm, title="Kalender Keberangkatan",
    )
    st = _pdf_styles()
    grp_note = " \u00b7 dikelompokkan per armada" if grouped else ""
    made = datetime.now().strftime('%d %b %Y %H:%M')
    story = [
        Paragraph("Kalender Keberangkatan", st["h1"]),
        Paragraph(f"RahazaTrans \u00b7 {period} \u00b7 {len(rows)} keberangkatan{grp_note} \u00b7 dibuat {made}", st["sub"]),
        Spacer(1, 5 * mm),
    ]
    if grouped:
        for name, grp in _group_rows(rows):
            story.append(Paragraph(f"\U0001F68C {name} \u2014 {len(grp)} keberangkatan", st["grp"]))
            story.append(_pdf_table(grp, st, include_vehicle=False))
            story.append(Spacer(1, 4 * mm))
        if not rows:
            story.append(Paragraph("Tidak ada keberangkatan pada periode ini.", st["sub"]))
    else:
        story.append(_pdf_table(rows, st, include_vehicle=True))
    doc.build(story)
    return buf.getvalue()


# --------------------------------------------------------------------------- XLSX
_HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
_HEAD_FILL = PatternFill("solid", fgColor="0058CC")
_THIN = Side(style="thin", color="D8DAE0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _xlsx_sheet(ws, title, subtitle, rows, include_vehicle=True):
    cols = COLS if include_vehicle else COLS_GROUPED
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="1C1C1E")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color="6B6B73")
    hr = 4
    for ci, col in enumerate(cols, start=1):
        c = ws.cell(row=hr, column=ci, value=col)
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    r = hr + 1
    for b in rows:
        for ci, v in enumerate(_values(b, include_vehicle), start=1):
            cell = ws.cell(row=r, column=ci, value=v)
            wrap_cols = (4, 7) if include_vehicle else (4, 6)
            cell.alignment = Alignment(vertical="center", wrap_text=(ci in wrap_cols))
            cell.border = _BORDER
        r += 1
    if not rows:
        ws.cell(row=r, column=1, value="Tidak ada keberangkatan.")
    widths = [12, 20, 20, 26, 20, 18, 28, 16, 12, 16] if include_vehicle else [12, 20, 20, 26, 18, 28, 16, 12, 16]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=hr, column=ci).column_letter].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


def _safe_sheet_name(name, used):
    s = re.sub(r"[\[\]:\*\?/\\]", " ", str(name or "Armada")).strip()[:28] or "Armada"
    base = s
    i = 2
    while s.lower() in used:
        s = f"{base[:26]} {i}"
        i += 1
    used.add(s.lower())
    return s


def calendar_xlsx(period: str, rows: list, grouped: bool = False) -> bytes:
    wb = Workbook()
    made = datetime.now().strftime('%d %b %Y %H:%M')
    if grouped:
        groups = _group_rows(rows)
        # Sheet ringkasan
        ws0 = wb.active
        ws0.title = "Ringkasan"
        ws0["A1"] = "Kalender Keberangkatan \u2014 Ringkasan per Armada"
        ws0["A1"].font = Font(bold=True, size=14, color="1C1C1E")
        ws0["A2"] = f"RahazaTrans \u00b7 {period} \u00b7 {len(rows)} keberangkatan \u00b7 dibuat {made}"
        ws0["A2"].font = Font(size=10, color="6B6B73")
        for ci, col in enumerate(["Armada", "Jumlah Keberangkatan"], start=1):
            c = ws0.cell(row=4, column=ci, value=col)
            c.font = _HEAD_FONT; c.fill = _HEAD_FILL; c.border = _BORDER
            c.alignment = Alignment(horizontal="center")
        rr = 5
        for name, grp in groups:
            ws0.cell(row=rr, column=1, value=name).border = _BORDER
            ws0.cell(row=rr, column=2, value=len(grp)).border = _BORDER
            rr += 1
        ws0.column_dimensions["A"].width = 30
        ws0.column_dimensions["B"].width = 22
        # Satu sheet per armada
        used = {"ringkasan"}
        if not groups:
            ws = wb.create_sheet("Kosong")
            _xlsx_sheet(ws, "Kalender Keberangkatan", f"{period} \u00b7 0 keberangkatan", [], include_vehicle=False)
        for name, grp in groups:
            ws = wb.create_sheet(_safe_sheet_name(name, used))
            _xlsx_sheet(ws, name, f"{period} \u00b7 {len(grp)} keberangkatan \u00b7 dibuat {made}", grp, include_vehicle=False)
    else:
        ws = wb.active
        ws.title = "Keberangkatan"
        _xlsx_sheet(ws, "Kalender Keberangkatan", f"RahazaTrans \u00b7 {period} \u00b7 {len(rows)} keberangkatan \u00b7 dibuat {made}", rows, include_vehicle=True)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
